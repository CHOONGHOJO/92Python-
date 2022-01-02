import datetime
from openpyxl import Workbook
wb = Workbook()
ws = wb.active

# A열의 너비를 5로 설정
ws.column_dimensions['A'].width = 18

ws['A1'] = datetime.datetime.today() # 오늘 날자
ws['A2'] = '=SUM(1,2,3)' # 1+2+3 = 6
ws['A3'] = '=AVERAGE(1,2,3)' # 1+2+3 = 6/3 = 2 (평균)

ws['A4'] = 10
ws['A5'] = 20
ws['A6'] = '=SUM(A4:A5)' # 30

wb.save("sample12_formula.xlsx")