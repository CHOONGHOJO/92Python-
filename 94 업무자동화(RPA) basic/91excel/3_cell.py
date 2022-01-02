from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "NadoSheet"

# A1 cell에 1 이라는 값을 입력
ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

print(ws["A1"]) # A1 cell의 정보 출력
print(ws["A1"].value) # A1 cell의 '값' 출력
print(ws["A10"].value) # 값이 없을 땐 'None' 출력

# column = A(1), B(2), C(3), ...
# row = 1, 2, 3, ...
print(ws.cell(column=1, row=1).value) # = ws["A1"].value
print(ws.cell(column=2, row=1).value) # = ws["B1"].value

c  = ws.cell(column=3, row=1, value=10) # ws["C1"].value = 10
print(c.value) # ws["C1"].value
ws["C1"] = 31
print(ws["C1"].value)

from random import *
index = 1
for x in range(1, 11): # 10 row
  for y in range(1, 11): # 10 column
    # ws.cell(column=y, row=x, value=int(str(y)+str(x))) # 0~100 사이의 숫자
    ws.cell(column=y, row=x, value=index) # 0~100 사이의 숫자
    index +=1




wb.save("sample.xlsx")
