from openpyxl import load_workbook
# wb = load_workbook("sample12_formula.xlsx")
# ws = wb.active

# # 수식 그대로 가져옴
# for row in ws.values:
#   for cell in row:
#     print(cell)


wb = load_workbook("sample12_formula.xlsx", data_only=True)
ws = wb.active

# 수식이 아닌 실제데이타를 가져옴
# evaluate 되지 않은 상태의 데이타는 None 이라고 표시되므로
#   sample12_formula.xlsx을 열었다가 save할 때 '저장'을 한 후 실행하면
#   실제데이타를 얻을 수 있다.
for row in ws.values:
  for cell in row:
    print(cell)

wb.save("sample13_formula_dataonly.xlsx")