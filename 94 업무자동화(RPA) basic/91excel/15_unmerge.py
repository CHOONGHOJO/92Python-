from openpyxl import load_workbook
wb = load_workbook("sample14_merge.xlsx")
ws = wb.active


# B2:D2 병합되어 있던  셀을 해제
ws.unmerge_cells("B2:D2") # B2 ~ D2 까지 해체


wb.save("sample15_unmerge.xlsx")