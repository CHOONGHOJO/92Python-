from openpyxl import Workbook
wb = Workbook()
ws = wb.active

# 병합하기
ws.merge_cells("B2:D2") # B2 ~ D2 까지 합침
ws["B2"].value = "Merged Cell"


wb.save("sample14_merge.xlsx")